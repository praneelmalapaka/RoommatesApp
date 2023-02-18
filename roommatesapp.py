from tkinter import *
from openpyxl import *
from openpyxl.worksheet.table import *
import datetime
import calendar
import smtplib
import webbrowser
from pandas import DataFrame
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
root = Tk()
root.geometry("1000x800")
root.title('Roommates App')
mths=["January", "February", "March", "April", "May", "June", "July", "August", "September", "October", "November", "December"]
dats=[31, 28, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31]
days=["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
lstname=[]
lstemail=[]
lstprice=[]
lstpswrd=[]
x = datetime.datetime.now()
global week
week=1
global co
co=0
global num
num=4
s=x.strftime("%x")[0]+x.strftime("%x")[1]
date=0
mth=0
global usrnm
usrnm=StringVar()
pswrd=StringVar()
#Function to initialise main screen which contains calendar GUI, the option to add expenditures and the option to remove expenditures
def mainscreen():
    bw=Workbook()
    bw=load_workbook(r"C:\Users\prane\Documents\School\Comp IA\verify.xlsx")
    sh=bw.active
    nms=sh["1"]
    psw=sh["2"]
    namesh=[]
    psswrd=[]    
    for i in nms:
        namesh.append(i.value)
    for j in psw:
        psswrd.append(j.value)
    if usrnm.get() not in namesh:
        top = Toplevel()
        Label(top, text=" ").pack()
        Label(top, text="Error! Data is incorrect!").pack()
        Label(top, text=" ").pack()
        Button(top, text="Ok", command=lambda: top.destroy()).pack()   
    elif str(psswrd[int(namesh.index(usrnm.get()))])==str(pswrd.get()) and (usrnm.get() in namesh):
        ppt = Toplevel()
        ppt.geometry("1000x800")
        Label(ppt, text="Red indicates the current date").grid(row=1,column=20)
        Label(ppt, text="Blue indicates payment deadlines").grid(row=2,column=20)
        Label(ppt, text="Green indicates that you have missed a payment deadline").grid(row=3,column=20)
        Label(ppt, text="Click on a date to add or remove an expediture").grid(row=5,column=20)
        #Function used to display graphical trends in expenditures
        def balance():
            topp=Toplevel()
            topp.geometry("600x900")
            book=Workbook()
            book=load_workbook(r"C:\Users\prane\Documents\School\Comp IA\dates.xlsx")
            work=book.active
            namebook=Workbook()
            namebook=load_workbook(r"C:\Users\prane\Documents\School\Comp IA\verify.xlsx")
            namesheet=namebook.active
            names=[]
            for t in namesheet["1"]:
                names.append(str(t.value))
            fulldata=[]
            tot=[]
            amount=[]
            indexcounter=0
            sumexpense=0
            for e in work["D"]:
                tot.append(str(e.value))
            for f in work["E"]:
                amount.append(int(f.value))
            genexpense=["Water bill", "Electricity bill", "Rent", "Miscellanious"]
            for k in genexpense:
                for r in tot: 
                    if r==k:
                        sumexpense=sumexpense+amount[indexcounter]
                    indexcounter=indexcounter+1
                fulldata.append(sumexpense)
                sumexpense=0
                indexcounter=0
            totb=[]
            amt=[]
            for x in work["C"]:
                totb.append(str(x.value))
            for m in names:
                for n in totb:
                    if m==n:
                        sumexpense=sumexpense+amount[indexcounter]
                    indexcounter=indexcounter+1
                amt.append(sumexpense)
                indexcounter=0
            
            twodarray=[]
            temparray=[]
            for z in range(len(names)):
                temparray.append(names[z])
                temparray.append(amt[z])
                twodarray.append(temparray)
                temparray=[]
            maximum=0
            h=0
            temp=0
            for maximum in range(len(twodarray)):
                h=maximum
                for current in range(maximum+1,len(twodarray)):
                    if twodarray[current][1]<twodarray[h][1]:
                        h=current
                temp=twodarray[h]
                twodarray[h]=twodarray[maximum]
                twodarray[maximum]=temp
            #The following code determines which user 
            statement="The user who spends the most amount of money is "+str(twodarray[0][0])+" with a total expenditure of $"+str(twodarray[0][1])
            Label(topp, text=statement).pack()
            completedata = {'Expenditure': ["Elec.","Misc.","Rent","Water"] ,'Amount': fulldata}
            df1 = DataFrame(completedata,columns=['Expenditure','Amount'])           
            #The folloing code block implements the bar graph         
            figure1 = plt.Figure(figsize=(10,1), dpi=100)
            ax1 = figure1.add_subplot(111)
            bar1 = FigureCanvasTkAgg(figure1, topp)
            bar1.get_tk_widget().pack(side=LEFT, fill=BOTH)
            df1 = df1[['Expenditure','Amount']].groupby('Expenditure').sum()
            df1.plot(kind='bar', legend=True, ax=ax1)
            ax1.set_title('Graph depicting expenditure split')          
        Button(ppt, text="Check balances", command=balance).grid(row=1, column=100)
        #Function to add expenditures
        def enter():
            dates=[]
            global date
            global mth
            global poi
            dates.append(date)
            dates.append(mth)
            brokwook=Workbook()
            brokwook=load_workbook(r"C:\Users\prane\Documents\School\Comp IA\dates.xlsx")
            hseet=brokwook.active
            poi=[]
            global num
            topp = Toplevel()
            topp.geometry("500x800")
            global track
            track=1
            wb=Workbook()
            wb=load_workbook(r"C:\Users\prane\Documents\School\Comp IA\verify.xlsx")
            ws=wb.active
            names=ws[1]
            workbook=Workbook()
            workbook=load_workbook(r"C:\Users\prane\Documents\School\Comp IA\cost.xlsx")
            sheet=workbook.active
            poi=[]
            nmes=[]
            for i in names:
                nmes.append(i.value)    
            lst=[]
            lst.append(usrnm.get())
            dates.append(usrnm.get())
            def clicked(value):
                global track
                lst.append(value)
                dates.append(value)
                track=track+1
            r = StringVar()
            num=len(nmes)
            amount_var=IntVar()
            v = StringVar()
            #The following code block lets users decide which type of expenditure
            genexpense=["Water bill", "Electricity bill", "Rent", "Miscellanious"]
            for i in range(num):
                Radiobutton(topp, text=genexpense[i], variable=v, value=genexpense[i]).pack()
            Button(topp, text="Submit", command=lambda: clicked(v.get())).pack()
            costy=IntVar()
            def cost():
                global track
                track=track+1
                global poi
                poi=[]
                coolio=costy.get()
                lst.append(coolio)
                dates.append(coolio/(num-1))
                #The following function helps add another expenditure
                def again():
                    if track==3:
                        for i in nmes:
                            poi=[]
                            if i!=lst[0]:                        
                                poi.append(i)
                                poi.append(lst[1])
                                poi.append(lst[2])
                                sheet.append(poi)
                                workbook.save(filename="cost.xlsx")
                            poi=[]
                        hseet.append(dates)
                        brokwook.save(filename="dates.xlsx")
                        topp.destroy()
                        enter()
                    else:
                        toppp = Toplevel()
                        def click():
                            toppp.destroy()
                            topp.destroy()
                            enter()
                        Label(toppp, text="Please fill all three to enter an expense").pack()
                        Button(toppp, text="Ok", command=click).pack()
                def done():
                    poi=[]
                    for i in nmes:
                        if i!=lst[0]:
                            poi.append(i)
                            poi.append(lst[1])
                            poi.append(int(lst[2])/(num-1))
                            sheet.append(poi)
                            workbook.save(filename="cost.xlsx")
                        poi=[]
                    hseet.append(dates)
                    brokwook.save(filename="dates.xlsx")
                    topp.destroy()            
                Button(topp, text="Another?", command=again).pack()
                Button(topp, text="Done?", command=done).pack()
            Entry(topp, textvariable=costy).pack()
            Button(topp, text="Submit", command=cost).pack()
        #Function to remove expenditures
        def remove():
            wb=Workbook()
            wb=load_workbook(r"C:\Users\prane\Documents\School\Comp IA\dates.xlsx")
            ws=wb.active
            bw=Workbook()
            bw=load_workbook(r"C:\Users\prane\Documents\School\Comp IA\verify.xlsx")
            sw=bw.active
            global date
            global mth
            topp=Toplevel()
            topp.geometry("500x800")
            nmese=sw["1"]
            nms=[]
            smn=[]
            snm=[]
            tpe=[]
            for e in nmese:
                nms.append(e.value)
            for j in ws["C"]:
                snm.append(str(j.value))
            for t in nms:
                if t in snm:
                    smn.append(t)
            q=StringVar()
            #The below code block provides users with payment options
            def pay(vue):
                s="A"+str(vue)
                ws[s]=None
                wb.save("dates.xlsx")
                Label(topp, text="Select your payment method").pack()
                Button(topp, text="Paypal", command=lambda: webbrowser.open("https://www.paypal.com/signin?returnUri=https%3A%2F%2Fwww.paypal.com%2Fmyaccount%2Ftransfer&state=%2F") ).pack()
                Button(topp, text="Paylah", command=lambda: webbrowser.open("https://internet-banking.dbs.com.sg/IB/Welcome")).pack()
                Button(topp, text="Grab Pay", command=lambda: webbrowser.open("https://www.grab.com/sg/pay/card/grabpay/")).pack()
            def typ(vlue):          
                cter=0
                tpe=[]
                s=""
                p=""
                o=0
                q=StringVar()               
                for g in ws["A"]:                    
                    cter=cter+1
                    if g.value!=None:
                        if int(g.value)==date:                            
                            if int(ws["B"+str(cter)].value)==mth+1:                                
                                s=""
                                s="C"+str(cter)                            
                            if str(ws["C"+str(cter)].value)==str(vlue):
                                p=""
                                o=0
                                o=cter
                                p="D"+str(o)                               
                                tpe.append(str(ws[p].value))               
                for d in range(len(tpe)):
                    Radiobutton(topp, text=tpe[d], variable=q, value=tpe[d]).pack()
                Button(topp, text="Submit", command=lambda: pay(o)).pack()
            def clicked(value):
                typ(value)
            f=StringVar()
            for i in range(len(smn)):
                Radiobutton(topp, text=smn[i], variable=f, value=smn[i]).pack()
            Button(topp, text="Submit", command=lambda: clicked(f.get())).pack()
        #The below codeblock depicts what output is to be made based on the date selected in the virtual calendar
        def chose(b,c):
            torkbook=load_workbook(r"C:\Users\prane\Documents\School\Comp IA\dates.xlsx")
            sht=torkbook.active
            tpop = Toplevel()
            global date
            global mth
            mth=c
            date=b
            counter=1
            s=""
            Label(tpop, text="The total payments due on this day are:").pack()
            for u in sht["A"]:
                if u.value!=None:
                    if u.value==date:
                        s=str(sht["C"+str(counter)].value)+" has added an expenditure of "+str(sht["E"+str(counter)].value)+" for "+str(sht["D"+str(counter)].value)
                        Label(tpop, text=s).pack()
                        s=""
                counter=int(counter)+1
            def entera():
                tpop.destroy()
                enter()
            def removea():
                tpop.destroy()
                remove()
            Button(tpop, text="Enter expenditure", command=entera).pack()
            Button(tpop, text="Remove expenditure", command=removea).pack()
        #Below code block manages the calendar responses for dates that have already passed
        def prev():
            prt = Toplevel()
            Label(prt, text="Kindly choose a date beyond today to enter an expenditure deadline.").pack()
            def ok():
                prt.destroy()
            Button(prt, text="Ok", command=ok).pack()
        def prv(b):
            torkbook=load_workbook(r"C:\Users\prane\Documents\School\Comp IA\dates.xlsx")
            sht=torkbook.active
            tpop = Toplevel()
            global date
            date=b
            counter=1
            s=""
            for u in sht["A"]:
                if u.value==date:
                    s="You have missed the deadline for paying back "+str(sht["C"+str(counter)].value)+" an amount of "+str(sht["E"+str(counter)].value)+" for "+str(sht["D"+str(counter)].value)
                    Label(tpop, text=s).pack()
                    s=""
                counter=int(counter)+1
        #Function gives the users option to either add or remove an expenditure
        def choose(a,b):
            global date
            global mth
            date=a
            mth=b
            prt = Toplevel()
            def entera():
                prt.destroy()
                enter()
            def removea():
                top.destroy()
                remove()
            Button(prt, text="Enter expenditure", command=entera).pack()
            Button(prt, text="Remove expenditure", command=removea).pack()
        
        def current(a):
            global week
            week=1
            torkbook=load_workbook(r"C:\Users\prane\Documents\School\Comp IA\dates.xlsx")
            sht=torkbook.active
            tpop = Toplevel()
            counter=1
            s=""
            today=0
            today=int(x.strftime("%x")[3]+x.strftime("%x")[4])
            flag=0
            #The below code block allows users to view deadlines
            Label(tpop, text="The total payments due on this day are:").pack()
            for u in sht["A"]:
                if u.value!=None:
                    if int(u.value)==a:
                        s=str(sht["C"+str(counter)].value)+" has added an expenditure of "+str(sht["E"+str(counter)].value)+" for "+str(sht["D"+str(counter)].value)
                        Label(tpop, text=s).pack()
                        s=""
                        flag=flag+1
                counter=int(counter)+1
            if flag==0:
                Label(tpop, text="There are no payments due today!").pack()
            Label(tpop, text="No expenditure deadlines can be added for today! Kindly choose a later date.").pack() 
        for i in range(1, dats[int(x.month)-1]+1):
            global week
            global co
            today=int(x.strftime("%x")[3]+x.strftime("%x")[4])
            qorkbook=Workbook()
            qorkbook=load_workbook(r"C:\Users\prane\Documents\School\Comp IA\dates.xlsx")
            seet=qorkbook.active
            dts=[]
            #The below code block allows for the virtual calendar to be implemented
            for p in seet["A"]:
                if p.value != None:
                    dts.append(int(p.value))
            countr=1
            d=""
            Label(ppt, text=mths[int(x.month)-1]).grid(row=4, column=4)
            if week-1==7:
                week=0
                co=co+1
            for t in dts:
                if i==t:
                    d="B"+str(countr)
                    if int(seet[d].value)==1:
                        if seet["C"+str(countr)].value!=usrnm.get():
                            if i>today:
                                Button(ppt, text=str(i), command=lambda j=i: chose(j,1), bg="Blue", padx=20, pady=20).grid(row=6+co, column=week)
                            else:
                                Button(ppt, text=str(i), command=lambda j=i: prv(j), bg="Green", padx=20, pady=20).grid(row=6+co, column=week)
                        else:
                            if i>today:
                                s=str(i)
                                Button(ppt, text=s, command=lambda j=i: choose(j,1) , padx=20, pady=20).grid(row=6+co, column=week)
                            else:
                                s=str(i)
                                Button(ppt, text=s, command=prev, padx=20, pady=20).grid(row=6+co, column=week)
                    else:
                        if i>today:
                            s=str(i)
                            Button(ppt, text=s, command=lambda j=i: choose(j,1) , padx=20, pady=20).grid(row=6+co, column=week)
                        else:
                            s=str(i)
                            Button(ppt, text=s, command=prev, padx=20, pady=20).grid(row=6+co, column=week)
                countr=countr+1
            if i==today:
                Button(ppt, text=str(i), command=lambda j=i: current(j), bg="red", padx=20, pady=20).grid(row=6+co, column=week)
            elif i not in dts:
                if i>today:
                    s=str(i)
                    Button(ppt, text=s, command=lambda j=i: choose(j,1) , padx=20, pady=20).grid(row=6+co, column=week)
                else:
                    s=str(i)
                    Button(ppt, text=s, command=prev, padx=20, pady=20).grid(row=6+co, column=week)
            week=week+1
        #The above code generates the calendar for the current month while the below code generates the calendar for the next month
        week=1
        co=0
        s=x.strftime("%x")[0]+x.strftime("%x")[1]
        k=int(x.month)
        if k==12:
            k=0
        Label(ppt, text=mths[k]).grid(row=14, column=4)
        for w in range(1, dats[k]+1):
            borkqook=Workbook()
            borkqook=load_workbook(r"C:\Users\prane\Documents\School\Comp IA\dates.xlsx")
            tees=borkqook.active
            dts=[]
            for p in seet["A"]:
                if p.value != None:
                    dts.append(int(p.value))
            countr=1
            d=""
            if week-1==7:
                week=0
                co=co+1
            for t in dts:
                if w==t:
                    d="B"+str(countr)
                    if int(tees[d].value)==2:
                        Button(ppt, text=str(w), command=lambda j=w: chose(j), bg="Blue", padx=20, pady=20).grid(row=15+co, column=week)
                    else:
                        s=str(w)
                        Button(ppt, text=s, command=lambda j=w: choose(j,2) , padx=20, pady=20).grid(row=15+co, column=week)
                countr=countr+1
            if w not in dts:
                s=str(w)
                Button(ppt, text=s, command=lambda j=w: choose(j,2) , padx=20, pady=20).grid(row=15+co, column=week)
            week=week+1
#Function to register users
def registration():
    wb=Workbook()
    ws=wb.active
    workbook=Workbook()
    sheet=workbook.active
    bookwork=Workbook()
    sheetwork=bookwork.active
    #Function to accept details during registration
    def submit():
        global name
        global num
        name=name_var.get()
        
        pswrd=pswrd_var.get()
        lstname.append(name)
        
        lstpswrd.append(pswrd)
        namedupl=lstname
        #Function that enables the same function to be repeated
        def again():
            top.destroy()
            registration()
        #Function that stops the registration process
        def done():
            top.destroy()
        ws.append(lstname)

        
        wb.save(filename="emails.xlsx")
        sheet.append(lstname)
        workbook.save(filename="balance.xlsx")
        sheetwork.append(lstname)
        sheetwork.append(lstpswrd)
        bookwork.save(filename="verify.xlsx")        
        num=num+1
        Button(top, text="Another?", command=again).pack()
        Button(top, text="Done?", command=done).pack()
    top=Toplevel()
    top.geometry("500x200")
    name_var=StringVar()
    email_var=StringVar()
    pswrd_var=StringVar()
    #The following code is for the login process
    Label(top, text="Enter name").pack()
    Entry(top, textvariable=name_var).pack()
    Label(top, text="Enter password").pack()
    Entry(top, textvariable=pswrd_var).pack()
    Button(top, text="Submit", command=submit).pack()
    Label(top, text="Enter details to register").pack()
Label(root, text="Welcome to the roommates app!").pack()
Label(root, text="Enter username").pack()
Entry(root, textvariable=usrnm).pack()
Label(root, text="Enter password").pack()
Entry(root, textvariable=pswrd).pack()
Button(root, text="Submit", command=mainscreen).pack()
Button(root, text="First time, click here", command=registration).pack()
root.mainloop()
